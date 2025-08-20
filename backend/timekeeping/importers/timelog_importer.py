# timekeeping/importers/timelog_importer.py
from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import datetime, time
from decimal import Decimal
from typing import Iterable, Optional, Tuple

from django.db import transaction
from django.utils.timezone import make_naive

from timekeeping.models import TimeLog, Holiday
from employees.models import Employee


@dataclass
class ImportOptions:
    employee_field: str = "employee_id"    # "employee_id" | "employee_number" | "email"
    date_format: str = "%Y-%m-%d"          # e.g. "2025-08-20"
    time_format: str = "%H:%M"             # e.g. "09:00"
    has_header: bool = True
    dry_run: bool = False                  # don't write to DB
    default_ot_hours: Decimal = Decimal("0.00")
    default_late_minutes: int = 0
    default_undertime_minutes: int = 0


@dataclass
class ImportResult:
    total_rows: int
    created: int
    updated: int
    skipped: int
    errors: list


REQUIRED_COLUMNS = {
    "employee_id|employee_number|email",
    "date",
}
# Optional columns we’ll ingest if present
OPTIONAL_COLUMNS = {
    "time_in", "time_out",
    "ot_hours", "late_minutes", "undertime_minutes",
    "is_rest_day", "is_absent",
    "holiday_code", "holiday_name",
}


def _to_bool(v) -> bool:
    if v is None:
        return False
    s = str(v).strip().lower()
    return s in {"1", "true", "yes", "y", "t"}


def _parse_time(value: Optional[str], time_format: str) -> Optional[time]:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, time):
        return value
    # Excel might give a datetime or time object—handle gracefully
    if isinstance(value, datetime):
        return value.time()
    try:
        return datetime.strptime(str(value).strip(), time_format).time()
    except Exception:
        # Try HH:MM fallback
        try:
            return datetime.strptime(str(value).strip(), "%H:%M").time()
        except Exception:
            raise


def _parse_decimal(value: Optional[str], default: Decimal) -> Decimal:
    if value is None or str(value).strip() == "":
        return default
    return Decimal(str(value))


def _parse_int(value: Optional[str], default: int) -> int:
    if value is None or str(value).strip() == "":
        return default
    return int(str(value))


def _get_employee(row: dict, employee_field: str) -> Employee:
    if employee_field == "employee_id":
        return Employee.objects.get(id=int(row.get("employee_id")))
    if employee_field == "employee_number":
        return Employee.objects.get(employee_number=str(row.get("employee_number")).strip())
    if employee_field == "email":
        return Employee.objects.get(email__iexact=str(row.get("email")).strip())
    raise ValueError("Unsupported employee_field")


def _get_holiday(row: dict) -> Optional[Holiday]:
    code = (row.get("holiday_code") or "").strip()
    name = (row.get("holiday_name") or "").strip()
    if code:
        try:
            return Holiday.objects.get(code=code)
        except Holiday.DoesNotExist:
            return None
    if name:
        try:
            return Holiday.objects.get(name__iexact=name)
        except Holiday.DoesNotExist:
            return None
    return None


def _normalize_row_keys(keys: Iterable[str]) -> set:
    return {str(k).strip() for k in keys}


def _is_excel(upload_name: str) -> bool:
    low = upload_name.lower()
    return low.endswith(".xlsx") or low.endswith(".xlsm") or low.endswith(".xls")


def _iter_rows_from_file(file_obj, filename: str, has_header: bool) -> Tuple[Iterable[dict], set]:
    """
    Yields rows as dicts with string keys. Returns (iterator, columns_set).
    Supports CSV and Excel. For Excel we use openpyxl (install it if needed).
    """
    if _is_excel(filename):
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            raise RuntimeError("openpyxl is required to read Excel files. pip install openpyxl")

        file_obj.seek(0)
        wb = openpyxl.load_workbook(file_obj, data_only=True, read_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = []
        if has_header:
            headers = [str(c).strip() if c is not None else "" for c in next(rows)]
        else:
            # If no header: generate generic headers
            first = next(rows)
            headers = [f"col_{i+1}" for i in range(len(first))]
            rows = (first,) + tuple(rows)
        columns = _normalize_row_keys(headers)

        def _gen():
            for r in rows:
                d = {headers[i]: r[i] if i < len(r) else None for i in range(len(headers))}
                yield d

        return _gen(), columns

    # CSV
    file_obj.seek(0)
    # Try to decode as UTF-8; if not, fallback to latin-1
    raw = file_obj.read()
    if isinstance(raw, bytes):
        try:
            text = raw.decode("utf-8")
        except UnicodeDecodeError:
            text = raw.decode("latin-1")
    else:
        text = raw

    sio = io.StringIO(text)
    if has_header:
        reader = csv.DictReader(sio)
        columns = _normalize_row_keys(reader.fieldnames or [])
        return reader, columns
    else:
        reader = csv.reader(sio)
        first = next(reader)
        headers = [f"col_{i+1}" for i in range(len(first))]
        columns = _normalize_row_keys(headers)

        def _gen():
            yield {headers[i]: first[i] for i in range(len(headers))}
            for r in reader:
                yield {headers[i]: (r[i] if i < len(r) else None) for i in range(len(headers))}

        return _gen(), columns


@transaction.atomic
def import_timelogs(
    file_obj,
    filename: str,
    options: Optional[ImportOptions] = None
) -> ImportResult:
    """
    Import timelogs from CSV/Excel. Upserts by (employee, date).
    """
    opts = options or ImportOptions()
    rows_iter, columns = _iter_rows_from_file(file_obj, filename, opts.has_header)

    # Validate minimal columns
    if not any(k in columns for k in ["employee_id", "employee_number", "email"]):
        raise ValueError("Missing one of required employee columns: employee_id | employee_number | email")
    if "date" not in columns:
        raise ValueError("Missing required column: date")

    total = created = updated = skipped = 0
    errors: list = []

    for row in rows_iter:
        total += 1
        try:
            # Map basic fields
            employee = _get_employee(row, opts.employee_field)

            # Date may be python date, excel serial, or string
            raw_date = row.get("date")
            if isinstance(raw_date, datetime):
                date_val = make_naive(raw_date).date()
            elif hasattr(raw_date, "date") and callable(raw_date.date):
                date_val = raw_date.date()
            else:
                date_val = datetime.strptime(str(raw_date).strip(), opts.date_format).date()

            time_in = _parse_time(row.get("time_in"), opts.time_format)
            time_out = _parse_time(row.get("time_out"), opts.time_format)

            ot_hours = _parse_decimal(row.get("ot_hours"), opts.default_ot_hours)
            late_minutes = _parse_int(row.get("late_minutes"), opts.default_late_minutes)
            undertime_minutes = _parse_int(row.get("undertime_minutes"), opts.default_undertime_minutes)
            is_rest_day = _to_bool(row.get("is_rest_day"))
            is_absent = _to_bool(row.get("is_absent"))
            holiday = _get_holiday(row)

            # If both time_in and time_out empty and not a rest day/holiday, you may consider absent
            if time_in is None and time_out is None and not is_rest_day and holiday is None:
                # Only set absent if not explicitly provided
                if "is_absent" not in row or str(row.get("is_absent")).strip() == "":
                    is_absent = True

            # Upsert by unique key (employee, date)
            obj, was_created = TimeLog.objects.update_or_create(
                employee=employee,
                date=date_val,
                defaults=dict(
                    time_in=time_in,
                    time_out=time_out,
                    ot_hours=ot_hours,
                    late_minutes=late_minutes,
                    undertime_minutes=undertime_minutes,
                    is_rest_day=is_rest_day,
                    is_absent=is_absent,
                    holiday=holiday,
                ),
            )

            if opts.dry_run:
                # Rollback at the end by raising to trigger atomic rollback
                pass

            created += 1 if was_created else 0
            updated += 0 if was_created else 1

        except Exception as e:
            skipped += 1
            errors.append(f"Row {total}: {e}")

    if opts.dry_run:
        # force rollback of everything in this atomic block
        transaction.set_rollback(True)

    return ImportResult(
        total_rows=total,
        created=created,
        updated=updated,
        skipped=skipped,
        errors=errors,
    )
